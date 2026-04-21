from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
import math


def style_excel(file_path):
    # Auto-adjust column widths and row heights
    wb = load_workbook(file_path)  # Opens the excel file and stores it in wb (workbook)
    ws = (
        wb.active
    )  # Gets the active sheet (first sheet) and stores it in ws (worksheet)

    # Auto-adjust column widths
    for col in ws.columns:  # Iterates through each column in the worksheet
        max_length = 0  # Initializes the maximum length of the column to 0 for tracking longest content
        column_letter = get_column_letter(
            col[0].column
        )  # Converts the column number to a letter (e.g. 1 -> A, 2 -> B, etc.)

        for cell in col:
            try:
                if cell.value:
                    # Calculate length considering newlines
                    cell_value = str(cell.value)  # Converts the cell value to a string
                    # For multiline cells, use the longest line
                    lines = cell_value.split(
                        "\n"
                    )  # Splits the cell value into lines based on newlines
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    max_length = max(max_length, max_line_length)
            except:
                pass

        # Set column width (add some padding, max 50 to avoid too wide columns)
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Auto-adjust row heights base on content
    for row in ws.iter_rows():
        for cell in row:
            try:
                if cell.value:
                    text = str(cell.value)
                    lines = text.split("\n")
                    col_letter = get_column_letter(cell.column)
                    col_width = ws.column_dimensions[col_letter].width or 10  # fallback
                    wrapped_lines = sum(
                        max(1, math.ceil(len(x) / col_width)) for x in lines
                    )

                    height = min(max(wrapped_lines * 15, 15), 200)
                    if (
                        ws.row_dimensions[cell.row].height is None
                        or ws.row_dimensions[cell.row].height < height
                    ):
                        ws.row_dimensions[cell.row].height = height
            except:
                pass

            # Enable text wrapping for better display
            cell.alignment = cell.alignment.copy(
                wrapText=True, vertical="top", horizontal="left"
            )

            # Make "Use Case" field and its value bold
            # NOTE: openpyxl Font() does not support `font_family=...`.
            # Use `name` to pick the font face (Excel will apply bold via bold=True).
            font_name = "Calibri"
            useCase_bold_font = Font(bold=True, size=14, name=font_name)
            useCase_fill_color = PatternFill(
                start_color="A5C9B1", end_color="A5C9B1", fill_type="solid"
            )
            testCase_fill_color = PatternFill(
                start_color="3A91B4", end_color="3A91B4", fill_type="solid"
            )
            testCase_bold_font = Font(bold=True, size=10, name=font_name)
            fill_color = PatternFill(
                start_color="92cddc", end_color="92cddc", fill_type="solid"
            )
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            cell.border = border

            if (
                cell.column == 1 and cell.value == "Use Case"
            ):  # Column 1 is "Field" column
                # Apply the same style to the first three cells in this row
                for col_idx in (1, 2, 3):
                    styled_cell = ws.cell(
                        row=cell.row, column=col_idx
                    )  # Gets the cell in the current row and column index
                    styled_cell.font = useCase_bold_font
                    styled_cell.fill = useCase_fill_color
            if (
                cell.column == 1
                and cell.value != "Use Case"
                and not isinstance(cell.value, (int, float))
            ):
                for col_idx in (1, 2, 3):
                    styled_cell = ws.cell(
                        row=cell.row, column=col_idx
                    )  # Gets the cell in the current row and column index
                    styled_cell.fill = fill_color
                    styled_cell.font = testCase_bold_font
            if cell.column == 1 and cell.value == "Test Case Title":
                for col_idx in (1, 2, 3):
                    styled_cell = ws.cell(
                        row=cell.row, column=col_idx
                    )  # Gets the cell in the current row and column index
                    styled_cell.fill = testCase_fill_color
    # Save the workbook
    wb.save(file_path)
    print(f"Excel file saved with auto-adjusted column widths and row heights")
